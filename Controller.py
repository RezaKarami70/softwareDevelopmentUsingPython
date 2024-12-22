import json, os, pyqtgraph as pg, numpy as np
import Model, View
from PyQt6 import QtWidgets
from PyQt6.QtCore import Qt
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter



class Controller():
    
    '''check is the path file exists or not and create it if not exists'''
    while True:
        try:
            with open("path.json", "r") as pathFile:
                pathDict = json.load(pathFile)
            d = pathDict["etabs_path"]
            d = pathDict["model_path"]
            d = pathDict["model_name"]
            d = pathDict["folder_path"]
            pathFile.close()
        except (FileNotFoundError, json.decoder.JSONDecodeError,KeyError):
            a_dict = {"etabs_path": "Import path to etabs", "model_path": "Import path to model", "model_name": "Import model name",
                   "folder_path": "Import path to saving folder", "previously_connected_to": ""}
            with open('path.json', 'w') as outfile:
                json.dump(a_dict, outfile)
        else:
            break
        
        
    def __init__(self):
        '''initionallaiz mode and view and show first window and actions'''
        self.view = View.view()
        self.model =  Model.Model()
        self.fillPathWindow()
        self.bottom_Action()
        
        
    def bottom_Action(self):
        '''all actions and signalls are here'''
        self.view.pathWindow.exitButton.pressed.connect(self.pathWindowExitButtonAction)
        self.view.pathWindow.closeButton.pressed.connect(self.pathWindowCloseButtonAction)
        self.view.pathWindow.runButton.pressed.connect(self.pathWindowRunButtonAction)
        self.view.pathWindow.ButtonEditEtabsPath.pressed.connect(self.ButtonEditEtabsPathAction)
        self.view.pathWindow.ButtonEditModelPath.pressed.connect(self.ButtonEditModelPathAction)
        self.view.pathWindow.ButtonEditSavePath.pressed.connect(self.ButtonEditSavePathAction)
        self.view.mainWindow.driftControlButton.pressed.connect(self.mainWindowDriftControlButtonAction)
        self.view.mainWindow.pathButton.pressed.connect(self.mainWindowpathButtonAction)
        self.view.mainWindow.closeButton.pressed.connect(self.mainWindowCloseButtenAction)
        self.view.mainWindow.connectButton.pressed.connect(self.mainWindowConnectButtonAction)
        self.view.driftControlWindow.closeButton.pressed.connect(self.driftControlWindowCloseACtions)
        self.view.driftControlWindow.listWidget.itemSelectionChanged.connect(self.sellectLoadCombAction)
        self.view.driftControlWindow.reortButton.pressed.connect(self.reportToExcel)
        self.view.driftControlWindow.my_signal.connect(self.mySlot)
        self.view.pathWindow.CatchButton.pressed.connect(self.getModelAndRun)
        self.view.driftControlWindow.selectoin.buttonClicked.connect(self.sellectLoadCombAction)
        
    def getModelAndRun(self):
        '''catch a model and run it user ressed Yes and breack if sellect cancel'''
        if self.wariningMessageWindow() == QtWidgets.QMessageBox.StandardButton.Cancel:
            return #cancel the running etabs and return 
        self.view.pathWindowForm.setDisabled(True)
        mesgBox = QtWidgets.QMessageBox()
        mesgBox.setText("please don't close etabs")
        mesgBox.exec()
        bl = self.model.getModel() # get a runing model
        if bl:
            self.view.pathWindowForm.setEnabled(True)
            self.view.pathWindowForm.close()
            self.view.mainWindowForm.show()
            self.view.mainWindow.lineEditPreConnect.setText(str(self.pathDict["previously_connected_to"])) # set the previous connection
            self.view.mainWindow.connectButton.setDisabled(True)
            self.view.mainWindow.driftControlButton.setEnabled(True)
        else:
            #erorr message
            self.view.pathWindowForm.setEnabled(True)
            mesgBox = QtWidgets.QMessageBox()
            mesgBox.setText("No running Etbas found")
            mesgBox.exec()
    
    def wariningMessageWindow(self):
        '''popup message for running etabs and retune it'''
        runMessage = QtWidgets.QMessageBox()
        runMessage.setStandardButtons(QtWidgets.QMessageBox.StandardButton.Ok | QtWidgets.QMessageBox.StandardButton.Cancel)
        runMessage.setText("Do you want to Run Etabs")
        ret = runMessage.exec()
        return ret

    def loadJson(self):
        '''load a json file for path handeling''' 
        with open("path.json", "r") as pathFile:
            self.pathDict = json.load(pathFile)
            path = self.pathDict
        return path
        
    
    def saveJson(self):
        '''save a dictionary as json file for path hanadling'''
        with open("path.json", "w") as pathFile:
            json.dump(self.pathDict, pathFile)
        return True
    
    def fillPathWindow(self):
        '''show first window and fill the path text lines'''
        self.view.pathWindow.lineEditEtabsPath.setText(self.pathDict["etabs_path"])
        self.view.pathWindow.lineEditModelPath.setText(self.pathDict["model_path"])
        self.view.pathWindow.lineEditModelName.setText(self.pathDict["model_name"])
        self.view.pathWindow.lineEditSavePath.setText(self.pathDict["folder_path"])
        self.view.pathWindowForm.show()
    
    def pathWindowExitButtonAction(self):
        '''close and exit the software'''
        self.view.pathWindowForm.close()
        
    def readJsonFile(self):
        '''update dict path for ready to save as jason file'''
        self.pathDict["etabs_path"] = self.view.pathWindow.lineEditEtabsPath.text()
        self.pathDict["model_path"] = self.view.pathWindow.lineEditModelPath.text()
        self.pathDict["model_name"] = self.view.pathWindow.lineEditModelName.text()
        self.pathDict["folder_path"] = self.view.pathWindow.lineEditSavePath.text()
        
    def pathWindowCloseButtonAction(self):
        '''save pathes and close path window and the software'''
        self.readJsonFile()
        self.saveJson()
        self.view.pathWindowForm.close()
    
    def pathWindowRunButtonAction(self):
        '''runing an open model or open and running a model from pathes imported whit user'''
        self.readJsonFile()
        self.view.pathWindowForm.close()
        #checking etabs software is exict or not
        if not os.path.exists(self.pathDict["etabs_path"] + os.sep + "ETABS.exe"):
            message = "Etabs.exe was not founded at " + self.pathDict["etabs_path"]
            self.WariningMessage(message)
            self.fillPathWindow()
            return
        #chicking model is exist ot not
        if not os.path.exists(self.pathDict["model_path"]):
            message = "model was not founded at " + self.pathDict["model_path"]
            self.WariningMessage(message)
            self.fillPathWindow()
            return
        #checking saving Excel report path is exist or not 
        if not os.path.exists(self.pathDict["folder_path"]):
            message = "Report path was not founded at " + self.pathDict["folder_path"]
            self.WariningMessage(message)
            self.fillPathWindow()
            return
        #if all the pathes are exist, save the json file going to main window
        self.saveJson()
        self.view.mainWindow.lineEditStatus.setText(str(self.pathDict["model_path"]))
        self.view.mainWindowForm.show()

    def WariningMessage(self, message):
        '''pop up message for directory'''
        mesgBox = QtWidgets.QMessageBox()
        mesgBox.setText(message)
        mesgBox.exec()


    def driftControlWindowCloseACtions(self):
        '''close drift window and emit signal for the main window'''
        self.view.driftControlWindow.emit_signal(self.update_label())#emit text to first window
        self.view.mainWindowForm.show()
        self.view.driftControlWindowForm.close()


    def mainWindowDriftControlButtonAction(self):
        '''close main window and show drift control window and fill list of combiations'''
        self.view.mainWindow.closeButton.setDisabled(True) 
        self.view.mainWindow.connectButton.setDisabled(True)
        self.view.mainWindow.pathButton.setDisabled(True)
        self.view.mainWindow.driftControlButton.setDisabled(True)
        self.fillCombolist() # fill list of comos
        self.view.mainWindow.closeButton.setEnabled(True)
        self.view.mainWindow.connectButton.setEnabled(True)
        self.view.mainWindow.pathButton.setEnabled(True)
        self.view.mainWindow.driftControlButton.setEnabled(True)
        self.view.driftControlWindow.line.setReadOnly = False
        self.view.driftControlWindowForm.show()
        self.view.mainWindowForm.hide()
        mesgBox = QtWidgets.QMessageBox()# wrining message
        mesgBox.setText("Pleas dont open Table in etabs")
        mesgBox.exec()
    
    def mainWindowCloseButtenAction(self):
        '''close the main window and close the software'''
        self.view.mainWindowForm.close()

    def mainWindowpathButtonAction(self):
        '''retun from main window to path window'''
        self.view.mainWindow.driftControlButton.setDisabled(True)
        self.view.mainWindow.lineEditPreDrift.setText("")
        self.view.driftControlWindow.listWidget.clear()
        self.view.driftControlWindow.tableWidget.clear()
        self.view.driftControlWindow.listWidget.scrollToTop()
        self.view.driftControlWindow.listWidget.clear()
        self.view.driftControlWindow.line.clear()
        self.view.driftControlWindow.plot.clear()
        self.view.mainWindowForm.close()
        self.fillPathWindow()
        
    
    def mainWindowConnectButtonAction(self):
        '''connecting to etabs and run analysis in main window update jason file for previous connection'''
        if self.wariningMessageWindow() == QtWidgets.QMessageBox.StandardButton.Cancel:
            return #return if the uuser not want to run the etabs 
        self.view.mainWindow.closeButton.setDisabled(True)
        self.view.mainWindow.connectButton.setDisabled(True)
        self.view.mainWindow.pathButton.setDisabled(True)
        self.view.mainWindow.driftControlButton.setDisabled(True)
        self.model.runModel()
        self.view.mainWindow.lineEditPreConnect.setText(str(self.pathDict["previously_connected_to"]))
        self.view.mainWindow.closeButton.setEnabled(True)
        self.view.mainWindow.connectButton.setEnabled(True)
        self.view.mainWindow.pathButton.setEnabled(True)
        self.view.mainWindow.driftControlButton.setEnabled(True)
        self.pathDict["previously_connected_to"] = self.pathDict["model_path"]
        with open("path.json", "w") as pathFile:# saving the current model as pre-connected model file
            json.dump(self.pathDict, pathFile)

    def selectDirectory(self):
        '''sellecting directory methode and return path'''
        self.dialog = QtWidgets.QFileDialog()
        self.folder_path = self.dialog.getExistingDirectory(None, "Select Folder")
        return self.folder_path

    def selectFile(self):
        '''sellecting EDB file methode and return model name and model path in a list'''
        self.dialog = QtWidgets.QFileDialog(caption="Choose File")
        file_filter = "ETABS Models (*.EDB)"
        self.filedir = self.dialog.getOpenFileNames(None, "Choose File", "", file_filter)
        try:
            self.name = self.filedir[0][0].split("/")[-1]
            return [self.name, self.filedir[0][0]]
        except(IndexError):
            return ["", ""]
        

    def ButtonEditEtabsPathAction(self):
        '''set directory to etabs path line edit'''
        path = self.selectDirectory()
        if not path == "":
            self.view.pathWindow.lineEditEtabsPath.setText(path)

    def ButtonEditModelPathAction(self):
        '''set directory to mode path llne edit'''
        path = self.selectFile()
        if not path[0] == "":
            self.view.pathWindow.lineEditModelName.setText(path[0])
            self.view.pathWindow.lineEditModelPath.setText(path[1])
        
    def ButtonEditSavePathAction(self):
        '''set diretory to report path line edit'''
        path = self.selectDirectory()
        self.view.pathWindow.lineEditSavePath.setText(path)
        
    def fillCombolist(self):        
        '''adding all load combinations to list in drift control window'''
        combosName = self.model.ComboName() #getings all load combinations names
        index = 0
        for comboName in combosName:
            item = QtWidgets.QListWidgetItem()
            self.view.driftControlWindow.listWidget.addItem(item) #adding item to list
            item = self.view.driftControlWindow.listWidget.item(index)
            index += 1
            item.setText(str(comboName))
            
    
    def fillComboTableAndLine(self):
        '''fill the table whit sellecet load combination in the drift control window list
        items by items'''
        self.view.driftControlWindow.tableWidget.setRowCount(len(self.storyslist))
        # seting width of 5 columns of tabls
        self.view.driftControlWindow.tableWidget.setColumnCount(5)
        self.view.driftControlWindow.tableWidget.setColumnWidth(0, 25)
        self.view.driftControlWindow.tableWidget.setColumnWidth(1, 65)
        self.view.driftControlWindow.tableWidget.setColumnWidth(2, 65)
        self.view.driftControlWindow.tableWidget.setColumnWidth(3, 65)
        self.view.driftControlWindow.tableWidget.setColumnWidth(4, 65)
        self.view.driftControlWindow.tableWidget.setHorizontalHeaderLabels(
            ["Sty", "XDisp(mm)", "YDisp(mm)","XDrift(mm)" ,"YDrift(mm)"]) #hedaings of table widgets
        self.maxDispX = 0 #maximum displacment in dir X for sellected combinations
        self.maxDispY = 0 #maximum displacment in dir Y for sellected combinations
        
        counter1 = 0 # rows of table
        for story in self.storyslist: # calcuate and adding story, X and Y Drinft and diplacement to table row by row
            try:
                XDrift = self.dfTable[( self.dfTable["Story"] == story)]['1000DriftX'].max().round(3)
            except(IndexError):
                XDrift = 0.0    #if etabs not return value for X drift in this story set 0.0 to it
            try:
                YDrift =  self.dfTable[( self.dfTable["Story"] == story)]['1000DriftY'].max().round(3)
            except(IndexError):
                YDrift = 0.0    #if etabs not return value for Y drift in this story set 0.0 to it
            try:
                Xdisp =  self.dfTable[( self.dfTable["Story"] == story)]['1000DispX'].max().round(3)
            except(IndexError):
                Xdisp = 0.0     #if etabs not return value for X dis in this story set 0.0 to it
            try:
                Ydisp =  self.dfTable[( self.dfTable["Story"] == story)]['1000DispY'].max().round(3)
            except(IndexError):
                Ydisp = 0.0     #if etabs not return value for Y dis in this story set 0.0 to it
                
            #calc max X and Y displacement
            if abs(self.maxDispX) < abs(Xdisp):
                self.maxDispX = Xdisp
            if abs(self.maxDispY) < abs(Ydisp):
                self.maxDispY = Ydisp
            
            #item create and set Alignments for adding to table
            item_story = QtWidgets.QTableWidgetItem(str(story))
            item_story.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            item_XDisp = QtWidgets.QTableWidgetItem(str("%.3f" % Xdisp))
            item_XDisp.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            item_YDisp = QtWidgets.QTableWidgetItem(str("%.3f" % Ydisp))
            item_YDisp.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            item_XDrift = QtWidgets.QTableWidgetItem(str("%.3f" % XDrift))
            item_XDrift.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            item_YDrift = QtWidgets.QTableWidgetItem(str("%.3f" % YDrift))
            item_YDrift.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

            items = [item_story, item_XDisp, item_YDisp, item_XDrift, item_YDrift]
            
            counter2 = 0 #columns of table
            for item in items:
                self.view.driftControlWindow.tableWidget.setItem(counter1, counter2, item) #adding each item to table
                counter2 += 1
            
            counter1 += 1
    
    def fillGraphForComboDis(self):
        '''fill the Graph for Displacement of sellected Laod'''
        self.view.driftControlWindow.plot.clear()
        ListX = self.dfTable['1000DispX'].to_list() #list of displacement X in milimeters
        ListY = self.dfTable['1000DispY'].to_list() #list of displacement Y in milimeters
        elev = self.dfTable['elevation'].to_list()  #ist of elevation in meters
        #list of points pass the ciritical canditions
        XDCritical = []
        XHCritical = []
        YDCritical = []
        YHCritical = []
        
        crt = self.calcdisThreshold() # threshold for displacements
        
        for i in range(len(ListX)):
            if abs(ListX[i]) > crt:
                XDCritical.append(ListX[i])
                XHCritical.append(elev[i]) # number of rejected points
            
        for i in range(len(ListY)):
            if abs(ListY[i]) > crt:
                YDCritical.append(ListY[i])
                YHCritical.append(elev[i]) # number of rejected points
                
        #threshold line        
        LimitLineXplus = np.array([crt, crt])
        LimitLineYplus = np.array([0, max(elev)])
        LimitLineXminus = np.array([-crt, -crt])
        LimitLineYminus = np.array([0, max(elev)])
        centerX = np.array([0,0])
        centerY = np.array([0,max(elev)])
        
        npXDCritical = np.array(XDCritical)
        npXHCritical = np.array(XHCritical)
        npYDCritical = np.array(YDCritical)
        npYHCritical = np.array(YHCritical)        


        # Create a line and ScatterPlotItem and add them to the PlotWidget
        self.view.driftControlWindow.plot.showGrid(x=True, y=True, alpha=0.5)
        self.view.driftControlWindow.plot.plot(ListY,elev, pen = "b", symbolPen ='b', symbol ='o', symbolSize = 8, name ="Disp Y")
        self.view.driftControlWindow.plot.plot(ListX,elev, pen = "r", symbolPen ='r', symbol ='o', symbolSize = 8, name ="Disp X")
        scatterX = pg.ScatterPlotItem(x=npXDCritical, y=npXHCritical, size=10, brush="m", name = "X-Dir Critical Disp") 
        self.view.driftControlWindow.plot.addItem(scatterX)
        scatterY = pg.ScatterPlotItem(x=npYDCritical, y=npYHCritical, size=10, brush="c",  name = "X-Dir Critical Disp")
        self.view.driftControlWindow.plot.addItem(scatterY)
        self.view.driftControlWindow.plot.plot(centerX,centerY, pen = "w")
        self.view.driftControlWindow.plot.setTitle(str(
            'Maxumim Disp For X-Dir=' + str("%.3f" %  max(ListX)) + ' and for Y-Dir=' + str("%.3f" % max(ListY)))) # Title
        if (len(XDCritical) + len(YDCritical)) > 0:
            LimitLineXplus = np.array([crt, crt])
            LimitLineYplus = np.array([0, max(elev)])
            LimitLineXminus = np.array([-crt, -crt])
            LimitLineYminus = np.array([0, max(elev)])
            self.view.driftControlWindow.plot.plot(LimitLineXplus,LimitLineYplus, pen = "y")
            self.view.driftControlWindow.plot.plot(LimitLineXminus,LimitLineYminus, pen = "y")
        
        

    def fillGraphForComboDrift(self):
        '''fill the Graph for Drift of sellected Laod'''
        self.view.driftControlWindow.plot.clear() #clear last plot
        ListX = self.dfTable['1000DriftX'].to_list() #list of X drift
        ListY = self.dfTable['1000DriftY'].to_list() #list of Y Drift
        elev = self.dfTable['elevation'].to_list() #list of elevetion
        #list of points pass the ciritical canditions
        XDriftCritical = []
        XHCritical = []
        YDriftCritical = []
        YHCritical = []
        crt = self.calcDriftThreshold()# threshold for Drift
        for i in range(len(ListX)):
            if ListX[i] > crt:
                XDriftCritical.append(ListX[i]) # number of rejected points
                XHCritical.append(elev[i])
            
        for i in range(len(ListY)):
            if ListY[i] > crt:
                YDriftCritical.append(ListY[i]) # number of rejected points
                YHCritical.append(elev[i])
                
        
        npXDriftCritical = np.array(XDriftCritical)
        npXHCritical = np.array(XHCritical)
        npYDriftCritical = np.array(YDriftCritical)
        npYHCritical = np.array(YHCritical)        


        # Create a ScatterPlotItem and a line plot and add it to the PlotWidget
        self.view.driftControlWindow.plot.showGrid(x=True, y=True, alpha=0.5)
        self.view.driftControlWindow.plot.plot(ListY,elev, pen = "b", symbolPen ='b', symbol ='o', symbolSize = 8, name ="Drift Y")
        self.view.driftControlWindow.plot.plot(ListX,elev, pen = "r", symbolPen ='r', symbol ='o', symbolSize = 8, name ="Drift X")
        scatterX = pg.ScatterPlotItem(x=npXDriftCritical, y=npXHCritical, size=10, brush="m", name = "X-Dir Critical Drift")
        self.view.driftControlWindow.plot.addItem(scatterX)
        scatterY = pg.ScatterPlotItem(x=npYDriftCritical, y=npYHCritical, size=10, brush="c",  name = "X-Dir Critical Drift")
        self.view.driftControlWindow.plot.addItem(scatterY)
        self.view.driftControlWindow.plot.setTitle(str('Maxumim Drift For X-Dir=' + str("%.3f" %  max(ListX)) + ' and for Y-Dir=' + str("%.3f" % max(ListY))))
        if (len(XDriftCritical) + len(YDriftCritical)) > 0:
            LimitLineX = np.array([crt, crt])
            LimitLineY = np.array([0, max(elev)])
            self.view.driftControlWindow.plot.plot(LimitLineX,LimitLineY, pen = "y")
        
    # action of sellected item in ist of control window change
    def sellectLoadCombAction(self):
        '''When a load is selected Graph and table is fll with this function'''
        try:
            self.dfTable = self.model.storyDispForCombo(self.view.driftControlWindow.listWidget.currentItem().text())
            self.storyslist = self.dfTable["Story"].to_list()
            if self.view.driftControlWindow.ra.isChecked(): #if radio button of Drift is checked
                self.fillGraphForComboDrift()  
                self.view.driftControlWindow.label_2.setText("Critcal Drift is " + str(self.calcDriftThreshold()) + "mm (Not true only fo cheking the app)")
            if self.view.driftControlWindow.rb.isChecked(): #if radio button of Displacement is checked
                self.fillGraphForComboDis()  
                self.view.driftControlWindow.label_2.setText("Critcal Displacment is " + str(self.calcdisThreshold()) + "mm (Not true only fo cheking the app)")
            self.fillComboTableAndLine()  

        except Exception as e:
            self.view.driftControlWindow.tableWidget.clear()
            self.view.driftControlWindow.plot.clear()
            mesgBox = QtWidgets.QMessageBox()
            mesgBox.setText("DisConneted from etabs, Close etabs and Connect again.")
            mesgBox.exec()


    def reportToExcel(self):
        '''create an Ecxcel file in selected directory with imported name'''
        #check name of excel file is import or not in a pop up window
        if self.view.driftControlWindow.line.text() == "":
            mesgBox = QtWidgets.QMessageBox()
            mesgBox.setText("Reort Name is Empty")
            mesgBox.exec()
            return
        
        try:        
            drift_threshold = self.calcDriftThreshold()
            #creat a excel workbook as object
            wb = Workbook()
            ws = wb.active
            ws.title = "Drift Analysis Report" #set the name of sheet
            #set the headings of table
            ws.cell(row=1, column=1, value="Story")
            ws.cell(row=1, column=2, value='DispX(mm)')
            ws.cell(row=1, column=3, value="DispY(mm)")
            ws.cell(row=1, column=4, value="DriftX(mm)")
            ws.cell(row=1, column=5, value="DriftY(mm)")

            rowIndex = 2 #rows in the workbook
            for story in self.storyslist:
                try:
                    XDrift = self.dfTable[( self.dfTable["Story"] == story)]['1000DriftX'].max().round(3)
                except(IndexError):
                    XDrift = 0.0    #set 0.0 for items not reported by etabs
                    
                try:
                    YDrift =  self.dfTable[( self.dfTable["Story"] == story)]['1000DriftY'].max().round(3)
                except(IndexError):
                    YDrift = 0.0    #set 0.0 for items not reported by etabs
                    
                try:
                    Xdisp =  self.dfTable[( self.dfTable["Story"] == story)]['1000DispX'].mean().round(3)
                except(IndexError):
                    Xdisp = 0.0    #set 0.0 for items not reported by etabs  
                    
                try:
                    Ydisp =  self.dfTable[( self.dfTable["Story"] == story)]['1000DispY'].mean().round(3)
                except(IndexError):
                    Ydisp = 0.0    #set 0.0 for items not reported by etabs  
                    
                cells = [story, Xdisp, Ydisp, XDrift, YDrift]                 
                colIndex = 1 #column index
                for cell in cells:
                    ws.cell(row=rowIndex, column=colIndex, value=cell)
                    colIndex += 1
                rowIndex += 1
                
                
            # Apply Conditional Formatting for Drift X and Drift Y Columns
            drift_x_col = 4  # Column for Drift X
            drift_y_col = 5  # Column for Drift Y

            # Add a ColorScaleRule for highlighting drift values
            color_scale = ColorScaleRule(
                start_type="num", start_value=drift_threshold*0.9, start_color="FFFFFFFF",  # White for low values
                mid_type="num", mid_value=drift_threshold *0.95, mid_color="FFFFFF00",  # Yellow
                end_type="num", end_value=drift_threshold, end_color="FFFF0000"  # Red for high values
            )

            # Apply the rule to Drift X and Drift Y
            ws.conditional_formatting.add(f"{get_column_letter(drift_x_col)}2:{get_column_letter(drift_x_col)}{len(self.storyslist)+1}", color_scale)
            ws.conditional_formatting.add(f"{get_column_letter(drift_y_col)}2:{get_column_letter(drift_y_col)}{len(self.storyslist)+1}", color_scale)

        
        except(AttributeError, ValueError, TypeError):
            mesgBox = QtWidgets.QMessageBox()
            mesgBox.setText("No selected Load or discanected From Etabs") # checking connection of etabs and selected load
            mesgBox.exec()
            return
        try:
            reportFileName = self.view.driftControlWindow.line.text()
            wb.save(self.pathDict["folder_path"] + "/" +reportFileName + ".xlsx") # saving excel file
        except(PermissionError): 
            mesgBox = QtWidgets.QMessageBox() #cheking file with same name is exist and open or not
            mesgBox.setText("There is an open file in " + self.pathDict["folder_path"] + "/" + reportFileName + " close the file or change the report name")
            mesgBox.exec()
            return
        except(FileNotFoundError):
            mesgBox = QtWidgets.QMessageBox()
            mesgBox.setText("There is no (" + self.pathDict["folder_path"] + " path") #cheking the directory
            mesgBox.exec()
            return
        mesgBox = QtWidgets.QMessageBox() #showing the path and saving condition
        mesgBox.setText("The report file is saved in (" + self.pathDict["folder_path"] + "/" + reportFileName + ").")
        mesgBox.exec()
        
        
    def calcDriftThreshold(self):
        '''calculates the Threshold for Drifts'''
        elev = self.dfTable['elevation'].to_list()
        if len(elev) -1 < 5:
            return 1000*0.025*max(elev)/400
        
        else:
            return 1000*0.02*max(elev)/400
        
    def calcdisThreshold(self):
        '''calculates the Threshold for Displacements'''
        elev = self.dfTable['elevation'].to_list()
        return 1000*0.0005*max(elev)

        

        
    def update_label(self):
        '''sent Max Displacement to the main window'''
        # Update the label with the received text
        try:
            preText = "for load " + self.view.driftControlWindow.listWidget.currentItem().text() + "Max Displacement for X=" + str("%.3f" % self.maxDispX) + "mm and for Y=" + str("%.3f" % self.maxDispY) + "mm"
            return preText
        except AttributeError:
            return ""
    
    def mySlot(self, text):
        '''sending signal to the main window'''
        self.view.mainWindow.lineEditPreDrift.setText(text)